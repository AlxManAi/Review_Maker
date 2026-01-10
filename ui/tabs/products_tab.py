"""
Products Tab - UI for managing product tasks with CSV import
"""
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, 
    QTableWidget, QTableWidgetItem, QHeaderView, QCheckBox,
    QMessageBox, QFileDialog, QLabel, QStyledItemDelegate, QLineEdit
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont
from ui.components.neon_button import NeonButton
from core.database import db
from core.models import ProductTask, Review
import time
import random

class TableItemDelegate(QStyledItemDelegate):
    """Custom delegate to fix text display during editing."""
    
    def createEditor(self, parent, option, index):
        editor = super().createEditor(parent, option, index)
        # Set minimum height for editor to 18px (шрифт 14px + 4px)
        if editor:
            editor.setMinimumHeight(26)
            font = QFont()
            font.setPointSize(14)
            editor.setFont(font)
        return editor


class ProductsTab(QWidget):
    """Tab for managing product tasks."""
    
    def __init__(self):
        super().__init__()
        self.current_period_id = None
        self.selected_product_ids = set()  # Сохраняем выбранные ID
        self._setup_ui()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Toolbar
        toolbar = QHBoxLayout()
        
        # Поиск
        search_label = QLabel("Поиск:")
        search_label.setStyleSheet("color: #ccc; margin-right: 5px;")
        toolbar.addWidget(search_label)
        
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Найти товар...")
        self.search_edit.setStyleSheet("""
            QLineEdit {
                background-color: #333;
                color: #fff;
                border: 1px solid #555;
                border-radius: 4px;
                padding: 5px;
                min-width: 200px;
            }
            QLineEdit:focus {
                border-color: #4a9eff;
            }
        """)
        self.search_edit.textChanged.connect(self.filter_products)
        toolbar.addWidget(self.search_edit)
        
        toolbar.addStretch()
        
        # Neon кнопки с умной активностью
        self.add_btn = NeonButton("Добавить", "primary")
        self.add_btn.clicked.connect(self.add_product)
        self.add_btn.setEnabled(False)
        toolbar.addWidget(self.add_btn)
        
        self.import_btn = NeonButton("Импорт", "secondary")
        self.import_btn.clicked.connect(self.import_products)
        self.import_btn.setEnabled(False)
        toolbar.addWidget(self.import_btn)
        
        self.export_btn = NeonButton("Экспорт", "secondary")
        self.export_btn.clicked.connect(self.export_products)
        self.export_btn.setEnabled(False)
        toolbar.addWidget(self.export_btn)
        
        self.parse_btn = NeonButton("Парсинг", "secondary")
        self.parse_btn.clicked.connect(self.parse_reviews)
        self.parse_btn.setEnabled(False)
        toolbar.addWidget(self.parse_btn)
        
        self.generate_btn = NeonButton("Генерация", "suggested")  # Рекомендуемое действие
        self.generate_btn.clicked.connect(self.generate_reviews)
        self.generate_btn.setEnabled(False)
        toolbar.addWidget(self.generate_btn)
        
        self.delete_btn = NeonButton("Удалить", "primary")  # Изменено на primary для единообразия
        self.delete_btn.clicked.connect(self.delete_products)
        self.delete_btn.setEnabled(False)
        toolbar.addWidget(self.delete_btn)

        layout.addLayout(toolbar)


        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["ID", "Название", "Кол-во", "URL товара", "Статус"])
        
        # Set column widths
        self.table.setColumnWidth(0, 50)   # ID - узкая
        self.table.setColumnWidth(2, 80)   # Кол-во - узкая
        self.table.setColumnWidth(4, 150)  # Статус - средняя
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)  # Название - средняя
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)  # URL - растягивается
        self.table.setColumnWidth(1, 200)  # Название - начальная ширина
        
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.verticalHeader().setVisible(False)
        
        # Set custom delegate to fix editing display
        self.table.setItemDelegate(TableItemDelegate())
        
        # Connect with error handling
        self.table.itemDoubleClicked.connect(self._safe_double_click)
        self.table.itemChanged.connect(self._safe_item_changed)
        layout.addWidget(self.table)
        
        # Review count validation label
        self.validation_label = QLabel("")
        self.validation_label = QLabel("")
        self.validation_label.setStyleSheet("padding: 5px;") # Removed bold
        layout.addWidget(self.validation_label)

    
    def _safe_double_click(self, item):
        """Safe wrapper for double click handler."""
        try:
            self.on_item_double_clicked(item)
        except Exception as e:
            import traceback
            error_msg = f"Ошибка при редактировании:\n{str(e)}\n\n{traceback.format_exc()}"
            print(error_msg)
            QMessageBox.critical(self, "Ошибка", error_msg)
    
    def _safe_item_changed(self, item):
        """Safe wrapper for item changed handler."""
        try:
            self.on_item_changed(item)
        except Exception as e:
            import traceback
            error_msg = f"Ошибка при сохранении:\n{str(e)}\n\n{traceback.format_exc()}"
            print(error_msg)
            QMessageBox.critical(self, "Ошибка", error_msg)

    
    def set_period(self, period_id):
        self.current_period_id = period_id
        self.add_btn.setEnabled(True)
        self.delete_btn.setEnabled(True)
        self.import_btn.setEnabled(True)
        self.export_btn.setEnabled(True)
        self.parse_btn.setEnabled(True)
        self.generate_btn.setEnabled(True)
        
        # Запуск умной активности
        self._update_smart_activity()
        
        self.load_products()
    
    def _update_smart_activity(self):
        """Обновление умной активности кнопок на основе контекста"""
        # Получаем количество товаров для анализа
        product_count = self.table.rowCount()
        
        # Сбрасываем все кнопки
        self.add_btn.stop_pulsing()
        self.parse_btn.stop_pulsing()
        self.generate_btn.stop_pulsing()
        
        # Логика умной активности:
        if product_count == 0:
            # Нет товаров - рекомендуем добавить
            self.add_btn.set_suggested(True)
        elif product_count > 0 and not self._has_parsed_reviews():
            # Есть товары, но нет отзывов - рекомендуем парсить
            self.parse_btn.set_suggested(True)
        elif self._has_parsed_reviews() and not self._has_generated_reviews():
            # Есть отзывы, но нет сгенерированных - рекомендуем генерировать
            self.generate_btn.set_suggested(True)
    
    def _has_parsed_reviews(self):
        """Проверка наличия спаршенных отзывов"""
        # Упрощенная проверка - можно улучшить
        return self.table.rowCount() > 0
    
    def _has_generated_reviews(self):
        """Проверка наличия сгенерированных отзывов"""
        # Упрощенная проверка - можно улучшить через БД
        return False  # Пока false, будем улучшать
    
    def refresh_validation(self):
        """Public method to refresh validation from external sources."""
        self.update_validation()


    
    def load_products(self):
        self.table.setRowCount(0)
        if not self.current_period_id:
            return
        
        # Устанавливаем принудительную высоту строк (Golden Air: 26px)
        self.table.verticalHeader().setDefaultSectionSize(33)
        
        self.table.blockSignals(True)  # Prevent itemChanged during load
        
        # Load products and extract data inside session
        products_data = []
        with db.get_session() as session:
            products = session.query(ProductTask).filter_by(period_id=self.current_period_id).all()

            product_ids = [p.id for p in products if isinstance(p.id, int)]
            generated_counts = {}
            try:
                from sqlalchemy import func
                if product_ids:
                    rows = (
                        session.query(Review.product_task_id, func.count(Review.id))
                        .filter(Review.product_task_id.in_(product_ids), Review.is_generated == True)
                        .group_by(Review.product_task_id)
                        .all()
                    )
                    generated_counts = {int(pid): int(cnt) for pid, cnt in rows if pid is not None}
            except Exception:
                generated_counts = {}

            for product in products:
                # Validate ID is integer
                if not isinstance(product.id, int):
                    print(f"ERROR: product.id is not integer: {product.id}, type: {type(product.id)}")
                    continue
                products_data.append({
                    'id': product.id,
                    'product_name': product.product_name,
                    'review_count': product.review_count,
                    'product_url': product.product_url,
                    'parse_status': product.parse_status,
                    'generated_count': int(generated_counts.get(product.id, 0) or 0)
                })
        
        # Add rows outside session using extracted data
        for data in products_data:
            self._add_product_row_from_data(data)
        
        self.table.blockSignals(False)
        
        # Update validation after loading
        self.update_validation()
    
    def update_validation(self):
        """Update review count validation label and generate button state."""
        if not self.current_period_id:
            self.validation_label.setText("")
            return
        
        with db.get_session() as session:
            # Get period's total_reviews_count
            from core.models import Period
            period = session.query(Period).get(self.current_period_id)
            if not period:
                return
            
            total_required = period.total_reviews_count
            
            # Calculate sum of product review counts
            from core.models import ProductTask
            products = session.query(ProductTask).filter_by(period_id=self.current_period_id).all()
            total_assigned = sum(p.review_count or 0 for p in products)
            
            # Update label
            if total_assigned == total_required:
                self.validation_label.setText(f"✓ Назначено отзывов: {total_assigned} / {total_required}")
                self.validation_label.setStyleSheet("padding: 5px; color: #50fa7b;") # Green
                self.generate_btn.setEnabled(True)
                self.generate_btn.setToolTip("")
            elif total_assigned < total_required:
                remaining = total_required - total_assigned
                self.validation_label.setText(f"⚠ Назначено отзывов: {total_assigned} / {total_required} (осталось: {remaining})")
                self.validation_label.setStyleSheet("padding: 5px; color: #ffb86c;") # Orange
                self.generate_btn.setEnabled(False)
                self.generate_btn.setToolTip(f"Назначьте ещё {remaining} отзывов для генерации")
            else:
                excess = total_assigned - total_required
                self.validation_label.setText(f"✗ Назначено отзывов: {total_assigned} / {total_required} (лишних: {excess})")
                self.validation_label.setStyleSheet("padding: 5px; color: #ff5555;") # Red
                self.generate_btn.setEnabled(False)
                self.generate_btn.setToolTip(f"Уберите {excess} отзывов - превышен лимит периода")

    
    def _add_product_row_from_data(self, data):
        """Add product row from data dictionary."""
        row = self.table.rowCount()
        self.table.insertRow(row)
        
        # ID (read-only but selectable)
        if not isinstance(data['id'], int):
            print(f"ERROR: data['id'] is not integer: {data['id']}, type: {type(data['id'])}")
            return
        id_item = QTableWidgetItem(str(data['id']))
        id_item.setFlags((id_item.flags() & ~Qt.ItemFlag.ItemIsEditable) | Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
        self.table.setItem(row, 0, id_item)
        
        # Editable fields (but need double-click to edit)
        name_item = QTableWidgetItem(data['product_name'])
        name_item.setFlags((name_item.flags() & ~Qt.ItemFlag.ItemIsEditable) | Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
        self.table.setItem(row, 1, name_item)
        
        count_item = QTableWidgetItem(str(data['review_count']) if data['review_count'] else "")
        count_item.setFlags((count_item.flags() & ~Qt.ItemFlag.ItemIsEditable) | Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
        self.table.setItem(row, 2, count_item)
        
        url_item = QTableWidgetItem(data['product_url'] or "")
        url_item.setFlags((url_item.flags() & ~Qt.ItemFlag.ItemIsEditable) | Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
        self.table.setItem(row, 3, url_item)
        
        # Статус парсинга
        status_text = data.get('parse_status', '')
        generated_count = int(data.get('generated_count', 0) or 0)
        required_count = int(data.get('review_count', 0) or 0)

        status_widget = QWidget()
        status_layout = QHBoxLayout(status_widget)
        status_layout.setContentsMargins(0, 0, 0, 0)
        status_layout.setSpacing(6)
        status_layout.setAlignment(Qt.AlignmentFlag.AlignVCenter)

        parsed_cb = QCheckBox()
        parsed_cb.setEnabled(False)
        parsed_cb.setChecked(status_text == 'success')
        parsed_cb.setFixedSize(18, 18)
        parsed_cb.setStyleSheet(
            """
            QCheckBox::indicator { width: 14px; height: 14px; border: 1px solid rgba(255,255,255,0.2); border-radius: 3px; background: rgba(255,255,255,0.05); }
            QCheckBox::indicator:checked { background: #4a9eff; border-color: rgba(74, 158, 255, 0.8); }
            """
        )
        if status_text == 'success':
            parsed_cb.setToolTip("Спарсенные отзывы: есть")
        elif status_text == 'failed':
            parsed_cb.setToolTip("Парсинг: ошибка")
        elif status_text == 'no_reviews':
            parsed_cb.setToolTip("Парсинг: отзывов не найдено")
        else:
            parsed_cb.setToolTip("Парсинг: не выполнялся")

        if required_count > 0:
            gen_label = QLabel(f"AI: {generated_count}/{required_count}")
        else:
            gen_label = QLabel(f"AI: {generated_count}")
        gen_label.setStyleSheet("color: #4a9eff; font-size: 12px; background: transparent;")
        gen_label.setToolTip(f"Сгенерировано отзывов: {generated_count} / {required_count}")

        status_layout.addWidget(parsed_cb, 0, Qt.AlignmentFlag.AlignVCenter)
        status_layout.addWidget(gen_label, 0, Qt.AlignmentFlag.AlignVCenter)
        status_layout.addStretch()

        status_item = QTableWidgetItem("")
        status_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
        self.table.setItem(row, 4, status_item)
        self.table.setCellWidget(row, 4, status_widget)
        
        # Включаем выделение строк для массовых операций
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.MultiSelection)

    
    def export_products(self):
        """Export products table to TXT file."""
        if not self.current_period_id:
            return
        
        file_path, _ = QFileDialog.getSaveFileName(self, "Экспорт товаров", "", "Text Files (*.txt)")
        if not file_path:
            return
        
        try:
            with db.get_session() as session:
                products = session.query(ProductTask).filter_by(period_id=self.current_period_id).all()
                
                with open(file_path, 'w', encoding='utf-8') as f:
                    # Write header
                    # f.write("Название;Количество;URL\n") 
                    # User likely wants raw data suitable for import, maybe without header? 
                    # Or with header but compatible. 
                    # Let's write simple format: Name;Count;URL
                    
                    for product in products:
                        line = f"{product.product_name};{product.review_count or ''};{product.product_url or ''}"
                        f.write(line + "\n")
            
            QMessageBox.information(self, "Успех", f"Экспортировано {len(products)} товаров")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка экспорта: {str(e)}")

    
    def on_item_double_clicked(self, item):
        """Double click = enable editing."""
        if item.column() in [0, 4]:  # Skip ID and checkbox
            return
        
        # Use Qt.ItemFlag directly to avoid recursion
        item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsEditable)
        self.table.editItem(item)

    
    def add_product(self):
        if not self.current_period_id:
            return
        
        with db.get_session() as session:
            # Проверяем на дубликаты перед добавлением
            existing_count = session.query(ProductTask).filter_by(
                period_id=self.current_period_id,
                product_name="Новый товар"
            ).count()
            
            if existing_count > 0:
                # Генерируем уникальное название
                base_name = "Новый товар"
                counter = 1
                while True:
                    new_name = f"{base_name} {counter}"
                    existing = session.query(ProductTask).filter_by(
                        period_id=self.current_period_id,
                        product_name=new_name
                    ).first()
                    if not existing:
                        break
                    counter += 1
                
                product_name = new_name
            else:
                product_name = "Новый товар"
            
            product = ProductTask(
                period_id=self.current_period_id,
                product_name=product_name,
                review_count=None,
                product_url=""
            )
            session.add(product)
            session.commit()
        
        # Reload all products to avoid detached instance issues
        self.load_products()
    
    def delete_products(self):
        """Delete selected products."""
        if not self.current_period_id:
            return
        
        # Get selected rows
        selected_rows = set(item.row() for item in self.table.selectedItems())
        if not selected_rows:
            QMessageBox.warning(self, "Предупреждение", "Выберите товары для удаления")
            return
        
        # Confirm deletion
        reply = QMessageBox.question(
            self, 
            "Подтверждение", 
            f"Удалить {len(selected_rows)} товар(ов)?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.No:
            return
        
        # Delete products
        with db.get_session() as session:
            for row in selected_rows:
                product_id = int(self.table.item(row, 0).text())
                product = session.query(ProductTask).get(product_id)
                if product:
                    session.delete(product)
            session.commit()
        
        self.load_products()
        QMessageBox.information(self, "Успех", f"Удалено {len(selected_rows)} товар(ов)")
    
    
    def import_products(self):
        """Import products from file (TXT/CSV or Excel)."""
        if not self.current_period_id:
            return
        
        file_path, _ = QFileDialog.getOpenFileName(self, "Импорт товаров", "", "Text Files (*.txt);;Excel Files (*.xlsx)")
        if not file_path:
            return
        
        try:
            products_to_add = []
            
            print(f"Импорт файла: {file_path}")
            
            if file_path.endswith('.xlsx'):
                from openpyxl import load_workbook
                wb = load_workbook(file_path)
                ws = wb.active
                # Skip header row
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                print(f"Найдено строк в Excel: {len(rows)}")
                
                for i, row in enumerate(rows, 2):
                    if not row or not row[0]: 
                        print(f"Пропуск строки {i}: пустая или без названия")
                        continue
                    
                    # Очистка данных
                    name = str(row[0]).strip() if row[0] else ""
                    count = None
                    url = ""
                    
                    # Проверяем остальные колонки
                    if len(row) > 1 and row[1]:
                        try:
                            count = int(row[1])
                        except (ValueError, TypeError):
                            print(f"Ошибка парсинга количества в строке {i}: {row[1]}")
                            pass
                    
                    if len(row) > 2 and row[2]:
                        url = str(row[2]).strip()
                    
                    if name:  # Только если есть название
                        products_to_add.append({
                            'name': name,
                            'count': count,
                            'url': url
                        })
                        print(f"Добавлен товар: {name} (кол-во: {count}, url: {url})")
                    else:
                        print(f"Пропуск строки {i}: нет названия")
                        
                print(f"Всего товаров для добавления: {len(products_to_add)}")
            else:
                # Text/CSV import
                with open(file_path, 'r', encoding='utf-8') as f:
                    for line in f:
                        line = line.strip()
                        if not line: continue
                        
                        # Try semicolon delimiter
                        parts = line.split(';')
                        name = parts[0].strip()
                        count = None
                        url = ""
                        
                        if len(parts) > 1:
                            # Try to parse 2nd field as count
                            try:
                                count = int(parts[1].strip())
                            except ValueError:
                                # If not int, maybe it's URL?
                                if parts[1].strip().startswith('http'):
                                    url = parts[1].strip()
                        
                        if len(parts) > 2:
                            url = parts[2].strip()
                            
                        products_to_add.append({
                            'name': name,
                            'count': count,
                            'url': url
                        })

            if not products_to_add:
                QMessageBox.warning(self, "Предупреждение", "Не найдено товаров для импорта")
                print("Ошибка: нет товаров для добавления")
                return

            print(f"Начинаем добавление {len(products_to_add)} товаров в базу данных")
            print(f"Текущий период ID: {self.current_period_id}")

            with db.get_session() as session:
                count_added = 0
                duplicates_found = []
                
                for i, p_data in enumerate(products_to_add, 1):
                    print(f"Обработка товара {i}: {p_data}")
                    
                    # Проверяем на дубликат по названию в текущем периоде
                    existing = session.query(ProductTask).filter_by(
                        period_id=self.current_period_id,
                        product_name=p_data['name']
                    ).first()
                    
                    if existing:
                        print(f"Найден дубликат: {p_data['name']}")
                        duplicates_found.append(p_data['name'])
                        continue
                    
                    product = ProductTask(
                        period_id=self.current_period_id,
                        product_name=p_data['name'],
                        review_count=p_data['count'],
                        product_url=p_data['url']
                    )
                    session.add(product)
                    count_added += 1
                    print(f"Товар добавлен в базу: {p_data['name']}")
                
                print(f"Коммит транзакции... Добавлено товаров: {count_added}")
                session.commit()
                print("Транзакция завершена")
                
                # Показываем предупреждение о дубликатах
                if duplicates_found:
                    QMessageBox.warning(
                        self, 
                        "Найдены дубликаты", 
                        f"Следующие товары уже существуют и не были добавлены:\n\n" + 
                        "\n".join(duplicates_found) + 
                        f"\n\nДобавлено уникальных товаров: {count_added}"
                    )
                else:
                    QMessageBox.information(self, "Успех", f"Импортировано {count_added} товаров")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка импорта: {str(e)}")

    
    def on_item_changed(self, item):
        row = item.row()
        col = item.column()
        
        try:
            product_id = int(self.table.item(row, 0).text())
        except (ValueError, AttributeError):
            return
        
        # Don't try to modify flags - causes recursion in PyQt6
        # Item will be made non-editable on next load
        
        try:
            with db.get_session() as session:
                product = session.query(ProductTask).get(product_id)
                if not product:
                    return
                
                if col == 1:
                    product.product_name = item.text()
                elif col == 2:
                    text = item.text().strip()
                    product.review_count = int(text) if text else None
                elif col == 3:
                    product.product_url = item.text()
                session.commit()
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка сохранения: {str(e)}")
            self.load_products()  # Reload to reset values
        
        # Update validation after changing review count
        self.update_validation()

    
    def parse_reviews(self):
        """Parse reviews from selected products to build training examples."""
        from core.logger import app_logger
        
        app_logger.info("=== НАЧАЛО ПАРСИНГА ОТЗЫВОВ ИЗ UI ===")
        
        try:
            if not self.current_period_id:
                app_logger.error("Нет текущего периода ID")
                return
            app_logger.info(f"Текущий период ID: {self.current_period_id}")
            
            # Получаем выделенные строки через selectedItems
            selected_rows = set()
            for item in self.table.selectedItems():
                selected_rows.add(item.row())
            
            selected_rows = list(selected_rows)
            
            if not selected_rows:
                app_logger.warning("Нет выбранных товаров")
                QMessageBox.warning(self, "Предупреждение", "Выберите товар(ы) для парсинга")
                return
            
            app_logger.info(f"Выбрано товаров: {len(selected_rows)}")
            
            # Собираем ID товаров
            product_ids = []
            for row in selected_rows:
                id_item = self.table.item(row, 0)
                if id_item:
                    product_ids.append(int(id_item.text()))
            
            # Создаем прогресс-диалог
            from ui.widgets.progress_dialog import ProgressDialog, ParseWorker
            progress_dialog = ProgressDialog("Парсинг отзывов", self)
            progress_dialog.set_title("Парсинг отзывов")
            progress_dialog.set_operation_details(f"Подготовка к парсингу {len(product_ids)} товаров...")
            
            # Создаем worker
            from core.parser_service import ParserService
            parser_service = ParserService()
            worker = ParseWorker(product_ids, parser_service)
            
            # Подключаем сигналы
            worker.progress_updated.connect(progress_dialog.update_progress)
            worker.finished.connect(lambda count, msg: self._on_parse_finished(count, msg, progress_dialog))
            worker.error.connect(lambda error: self._on_parse_error(error, progress_dialog))
            
            # Запускаем
            worker.start()
            progress_dialog.exec()
            
        except Exception as e:
            app_logger.error(f"Ошибка в parse_reviews: {e}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при парсинге: {str(e)}")
    
    def _on_parse_finished(self, count, message, dialog):
        """Обработка завершения парсинга"""
        dialog.finish(message)
        # Обновляем таблицу
        self.load_products()
    
    def _on_parse_error(self, error, dialog):
        """Обработка ошибки парсинга"""
        dialog.finish(f"Ошибка: {error}")
        QMessageBox.critical(dialog, "Ошибка парсинга", error)
    
    def show_parsed_reviews(self, product_ids):
        """Показать диалог для просмотра и утверждения спарсенных отзывов."""
        from ui.dialogs.parsed_reviews_dialog import ParsedReviewsDialog
        dialog = ParsedReviewsDialog(product_ids, parent=self)
        dialog.exec()
    
    def generate_reviews(self):
        """Генерация отзывов с помощью AI."""
        # Получить выбранные товары
        selected_rows = set(item.row() for item in self.table.selectedItems())
        if not selected_rows:
            QMessageBox.warning(self, "Предупреждение", "Выберите товар(ы)")
            return
        
        product_ids = []
        for row in selected_rows:
            try:
                id_item = self.table.item(row, 0)
                if id_item:
                    product_ids.append(int(id_item.text()))
            except ValueError:
                continue
                
        if not product_ids:
            return
            
        # Проверка (опционально)
        from core.database import db
        from core.models import ProductTask
        
        with db.get_session() as session:
            count = session.query(ProductTask).filter(ProductTask.id.in_(product_ids)).count()
            if count == 0:
                 QMessageBox.warning(self, "Ошибка", "Товары не найдены")
                 return
        
        # Открыть диалог генерации
        from ui.dialogs.generate_dialog import GenerateDialog
        dialog = GenerateDialog(parent=self, product_ids=product_ids)
        if dialog.exec():
            # Обновить валидацию после генерации
            self.load_products()
    
    def delete_product_reviews(self):
        """Удалить отзывы для выбранных товаров (только не принятые)."""
        if not self.current_period_id:
            QMessageBox.warning(self, "Предупреждение", "Выберите период")
            return
        
        selected_rows = set(item.row() for item in self.table.selectedItems())
        if not selected_rows:
            QMessageBox.warning(self, "Предупреждение", "Выберите товары для удаления отзывов")
            return
        
        # Собираем ID товаров
        product_ids = []
        for row in selected_rows:
            product_id = int(self.table.item(row, 0).text())
            product_ids.append(product_id)
        
        # Проверяем, есть ли принятые отзывы
        with db.get_session() as session:
            from core.models import Review
            from sqlalchemy import func
            
            # Проверяем наличие принятых отзывов
            approved_count = session.query(func.count(Review.id)).filter(
                Review.product_task_id.in_(product_ids),
                Review.is_approved == True
            ).scalar()
            
            if approved_count > 0:
                QMessageBox.warning(
                    self, 
                    "Ошибка", 
                    f"Нельзя удалить отзывы! У {approved_count} отзывов стоит флаг 'Принято'.\n\n"
                    "Удалять отзывы можно только до их принятия."
                )
                return
            
            # Удаляем не принятые отзывы
            deleted_count = session.query(Review).filter(
                Review.product_task_id.in_(product_ids),
                Review.is_approved == False
            ).delete()
            
            session.commit()
            
            if deleted_count > 0:
                QMessageBox.information(
                    self, 
                    "Успех", 
                    f"Удалено {deleted_count} отзывов для {len(product_ids)} товаров"
                )
            else:
                QMessageBox.information(
                    self, 
                    "Информация", 
                    "Нет отзывов для удаления (только принятые отзывы)"
                )

    def filter_products(self, text):
        """Фильтрация товаров по поисковому тексту."""
        search_text = text.lower().strip()
        
        # Сохраняем выбранные ID перед фильтрацией
        selected_rows = self.table.selectionModel().selectedRows()
        self.selected_product_ids = set()
        for row in selected_rows:
            id_item = self.table.item(row.row(), 0)
            if id_item:
                try:
                    self.selected_product_ids.add(int(id_item.text()))
                except ValueError:
                    pass
        
        for row in range(self.table.rowCount()):
            show_row = False
            
            # Ищем по всем колонкам
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item and search_text in item.text().lower():
                    show_row = True
                    break
            
            self.table.setRowHidden(row, not show_row)
            
            # Восстанавливаем выделение если строка видима и была выбрана
            if show_row:
                id_item = self.table.item(row, 0)
                if id_item:
                    try:
                        product_id = int(id_item.text())
                        if product_id in self.selected_product_ids:
                            self.table.selectRow(row)
                    except ValueError:
                        pass


