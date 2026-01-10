"""
Calendar Reviews Tab - Календарный просмотр отзывов
"""
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QScrollArea,
    QLabel, QMessageBox, QCheckBox,
    QTextEdit, QFrame, QGridLayout, QLineEdit, QPushButton
)
from PyQt6.QtCore import Qt, pyqtSignal
from datetime import datetime, timedelta
from ui.components.neon_button import NeonButton
import random


class ReviewCard(QFrame):
    """Карточка отзыва."""
    
    approved_changed = pyqtSignal(int, bool)  # review_id, is_approved
    used_changed = pyqtSignal(int, bool)  # review_id, is_used
    
    def __init__(self, review, view_mode='dashboard', parent=None):
        super().__init__(parent)
        self.review_id = review.id
        self.view_mode = view_mode
        self.product_name = review.product_name or ""
        self.author = review.author or ""
        self.content = review.content or ""
        self.init_ui(review)
    
    def init_ui(self, review):
        
        # Базовая стилизация карточки
        self.setFrameStyle(QFrame.Shape.NoFrame)
        
        # Unified Typography & Colors
        font_main = "font-family: 'Inter', sans-serif;"
        
        if self.view_mode == "minimalist":
            bg_style = "background: transparent; border: none;"
            hover_style = "background: rgba(74, 158, 255, 0.05); border-radius: 4px;"
            margins = (2, 4, 2, 4)
            spacing = 0
        elif self.view_mode == "kanban":
            # Kanban: Glassmorphism tile
            bg_style = "background: rgba(30, 31, 48, 0.4); border: 1px solid rgba(255, 255, 255, 0.07); border-radius: 10px;"
            hover_style = "background: rgba(74, 158, 255, 0.03); border-color: rgba(74, 158, 255, 0.3);"
            margins = (8, 10, 8, 10)
            spacing = 4
        else: # dashboard
            # Dashboard: Light Token
            bg_style = "background: rgba(45, 45, 69, 0.4); border: 1px solid rgba(255, 255, 255, 0.05); border-radius: 8px;"
            hover_style = "background: rgba(74, 158, 255, 0.05); border-color: rgba(74, 158, 255, 0.3);"
            margins = (4, 4, 4, 4)
            spacing = 2

        self.setStyleSheet(f"""
            ReviewCard {{
                {bg_style}
                {font_main}
            }}
            ReviewCard:hover {{
                {hover_style}
            }}
            QLabel {{
                background: transparent;
            }}
        """)
        
        layout = QVBoxLayout()
        layout.setContentsMargins(*margins)
        layout.setSpacing(spacing)
        
        # Header layout
        header = QHBoxLayout()
        header.setSpacing(4)
        
        # Brute-force Font Standardization (12px for all modes)
        prod_text = review.product_name or "Товар"
        if self.view_mode == "dashboard":
            if len(prod_text) > 18: prod_text = prod_text[:15] + "..."
        elif self.view_mode == "kanban":
            if len(prod_text) > 35: prod_text = prod_text[:32] + "..."
        else:  # minimalist
            if len(prod_text) > 40: prod_text = prod_text[:37] + "..."
            
        product_label = QLabel(prod_text)
        f = product_label.font()
        f.setPointSize(11)  # Unified size
        f.setWeight(600)    # Consistent weight
        product_label.setFont(f)
        product_label.setStyleSheet("color: #4A90E2; background: transparent;")
        product_label.setObjectName("ProductNameLabel")
            
        header.addWidget(product_label)
        
        header.addStretch()
        
        # Icons/Controls (разные для разных режимов)
        if self.view_mode == "dashboard":
            # Dashboard: Tiny perfect squares
            self.approved_checkbox = QCheckBox()
            self.used_checkbox = QCheckBox()
            
            # Subtly style indicators to be perfect squares
            check_css = """
                QCheckBox::indicator { width: 18px; height: 18px; border: 1px solid rgba(255,255,255,0.1); border-radius: 2px; background: rgba(255,255,255,0.05); }
                QCheckBox::indicator:checked { background: #4a9eff; }
            """
            self.approved_checkbox.setStyleSheet(check_css)
            self.used_checkbox.setStyleSheet(check_css)
            self.approved_checkbox.setFixedSize(24, 24) # Forced 24x24px
            self.used_checkbox.setFixedSize(24, 24) # Forced 24x24px
            
            self.approved_checkbox.setChecked(review.is_approved)
            self.approved_checkbox.toggled.connect(lambda v: self.approved_changed.emit(review.id, v))
            self.used_checkbox.setChecked(review.is_used)
            self.used_checkbox.toggled.connect(lambda v: self.used_changed.emit(review.id, v))
            
            self.open_btn = QPushButton("📝")
            self.open_btn.setObjectName("IconButton")
            self.open_btn.setFixedSize(24, 24) # Forced 24x24px
            
            header.addWidget(self.approved_checkbox)
            header.addWidget(self.used_checkbox)
            header.addWidget(self.open_btn)
            
        elif self.view_mode == "kanban":
            # Kanban: Standard perfect squares
            self.approved_checkbox = QCheckBox("✅")
            self.used_checkbox = QCheckBox("📦")
            
            checkbox_style = """
                QCheckBox { color: #888; font-size: 11px; spacing: 4px; }
                QCheckBox::indicator { width: 18px; height: 18px; border: 1px solid rgba(255,255,255,0.2); border-radius: 4px; background: rgba(255,255,255,0.05); }
                QCheckBox::indicator:checked { background: #4a9eff; border-color: #4a9eff; }
            """
            self.approved_checkbox.setStyleSheet(checkbox_style)
            self.used_checkbox.setStyleSheet(checkbox_style)
            # Ensure fixed size for alignment
            self.approved_checkbox.setFixedSize(24, 24) # Forced 24x24px
            self.used_checkbox.setFixedSize(24, 24) # Forced 24x24px
            
            self.approved_checkbox.setChecked(review.is_approved)
            self.approved_checkbox.toggled.connect(lambda v: self.approved_changed.emit(review.id, v))
            self.used_checkbox.setChecked(review.is_used)
            self.used_checkbox.toggled.connect(lambda v: self.used_changed.emit(review.id, v))
            
            self.open_btn = QPushButton("📝")
            self.open_btn.setObjectName("IconButton")
            self.open_btn.setFixedSize(24, 24) # Forced 24x24px
            
            header.addWidget(self.approved_checkbox)
            header.addWidget(self.used_checkbox)
            header.addWidget(self.open_btn)
            
        else:  # minimalist
            # Только кнопка редактирования
            self.approved_checkbox = QCheckBox()
            self.approved_checkbox.setText("")
            self.approved_checkbox.setChecked(review.is_approved)
            self.approved_checkbox.toggled.connect(lambda v: self.approved_changed.emit(review.id, v))
            
            self.used_checkbox = QCheckBox()
            self.used_checkbox.setText("")
            self.used_checkbox.setChecked(review.is_used)
            self.used_checkbox.toggled.connect(lambda v: self.used_changed.emit(review.id, v))
            
            # Кнопка редактирования
            self.open_btn = QPushButton("📝")
            self.open_btn.setObjectName("IconButton")
            self.open_btn.setFixedSize(24, 24)
            
            header.addWidget(self.open_btn)
        layout.addLayout(header)
        
        # Content display для разных режимов
        if self.view_mode == "kanban":
            # В канбане: автор, рейтинг и короткий текст
            author_rating = QLabel(f"👤 {review.author or 'Аноним'} | ⭐ {review.rating or 5}")
            author_rating.setStyleSheet("color: #888; font-size: 9px; padding: 1px 3px;")
            layout.addWidget(author_rating)
            
            # Короткий текст отзыва
            content_label = QLabel(review.content or "")
            content_label.setWordWrap(True)
            content_text = review.content or ""
            if len(content_text) > 80:
                content_text = content_text[:77] + "..."
            content_label.setText(content_text)
            content_label.setStyleSheet("color: #cdd6f4; font-size: 10px; padding: 2px 4px; line-height: 1.2;")
            layout.addWidget(content_label)
        elif self.view_mode == "dashboard":
            # В дашборде только товар и иконки, контент скрыт
            pass
        else: # minimalist
            # В минимализме карточки либо не показываются, либо максимально скрыты
            # Для полноты, если они рендерятся, скрываем всё
            self.hide()
        
        self.setLayout(layout)
        if self.view_mode == "kanban":
            self.setToolTip(review.content or "")

    
class DayContainer(QFrame):
    """Контейнер дня (вертикальный список дней)."""
    
    def __init__(self, date, view_mode='dashboard', parent=None):
        super().__init__(parent)
        self.date = date
        self.view_mode = view_mode
        self.setAcceptDrops(True)
        self.init_ui()
    
    def init_ui(self):
        if self.view_mode == "minimalist":
            border_style = "border-bottom: 2px solid rgba(255, 255, 255, 0.05);"
            bg_color = "transparent"
            margins = (4, 15, 4, 15)
            self.setMinimumHeight(28)
        elif self.view_mode == "dashboard":
            border_style = "border: 1px solid rgba(255, 255, 255, 0.08); border-radius: 12px;"
            bg_color = "rgba(17, 17, 27, 0.5)"
            margins = (10, 10, 10, 10)
            self.setMinimumHeight(160) # Increased height for desktop app
        else: # kanban
            border_style = "border: 1px solid rgba(255, 255, 255, 0.12); border-radius: 16px;"
            bg_color = "rgba(30, 30, 46, 0.5)"
            margins = (12, 12, 12, 12)
            self.setMinimumHeight(180)

        self.setStyleSheet(f"""
            DayContainer {{
                background-color: {bg_color};
                {border_style}
                margin: 0px;
            }}
            DayContainer:hover {{
                border-color: rgba(74, 158, 255, 0.6);
                background-color: rgba(74, 158, 255, 0.03);
            }}
        """)
        self.layout = QVBoxLayout()
        self.layout.setContentsMargins(*margins)
        self.layout.setSpacing(4)
        
        # Заголовок дня
        title_layout = QHBoxLayout()
        title_layout.setContentsMargins(0, 0, 0, 0)
        title_layout.setSpacing(6)
        
        # Русские дни недели
        russian_days = {
            'Mon': 'Пн', 'Tue': 'Вт', 'Wed': 'Ср', 'Thu': 'Чт',
            'Fri': 'Пт', 'Sat': 'Сб', 'Sun': 'Вс'
        }
        
        if self.view_mode == "dashboard":
            date_str = self.date.strftime("%d.%m")
        elif self.view_mode == "kanban":
            day_name = russian_days.get(self.date.strftime("%a"), self.date.strftime("%a"))
            date_str = self.date.strftime(f"%d.%m ({day_name})")
        else:  # minimalist
            date_str = self.date.strftime("%d.%m")
            
        self.date_label = QLabel(date_str)
        # Unified Header Font
        f = self.date_label.font()
        f.setPointSize(13)
        f.setBold(True)
        self.date_label.setFont(f)
        self.date_label.setMinimumHeight(22)
        self.date_label.setStyleSheet("color: #4a9eff; background: transparent; padding: 2px 0px;")
        title_layout.addWidget(self.date_label)
        
        if self.view_mode == "dashboard":
            title_layout.addStretch()
            self.stats_pill = QLabel(".")
            self.stats_pill.setMinimumHeight(18)
            self.stats_pill.setStyleSheet("""
                background-color: rgba(74, 158, 255, 0.06);
                color: #4a9eff;
                padding: 2px 8px;
                border-radius: 6px;
                font-size: 9px;
                border: 1px solid rgba(74, 158, 255, 0.1);
            """)
            title_layout.addWidget(self.stats_pill)
        elif self.view_mode == "kanban":
            title_layout.addStretch()
        else:  # minimalist
            # В минимализме показываем только статистику дня
            title_layout.addStretch()
            stats_text = f"✅ 0 | 📦 0 | 📝 0"  
            
            self.stats_label = QLabel(stats_text)
            self.stats_label.setStyleSheet("""
                color: #4a9eff; 
                font-size: 10px; 
                padding: 2px 8px; 
                background: rgba(74, 158, 255, 0.06); 
                border-radius: 10px;
                border: 1px solid rgba(74, 158, 255, 0.1);
            """)
            title_layout.addWidget(self.stats_label)
            
        self.layout.addLayout(title_layout)
        
        # Область для карточек отзывов
        self.reviews_area = QVBoxLayout()
        self.reviews_area.setSpacing(4 if self.view_mode == "dashboard" else 8)
        self.layout.addLayout(self.reviews_area)
        
        # Spacer inside grid cell to top-align content
        if self.view_mode == "dashboard":
            self.layout.addStretch()
            
        self.setLayout(self.layout)
    
    def add_review(self, review_card):
        """Добавить карточку отзыва в контейнер."""
        self.reviews_area.addWidget(review_card)

    def update_stats(self, approved_count, used_count, total_count):
        """Обновить статистику для всех режимов"""
        if hasattr(self, 'stats_pill'):
            # Для Dashboard
            self.stats_pill.setText(f"✅ {approved_count}/{total_count} | 📦 {used_count}")
            # Цветовая индикация прогресса
            if total_count > 0 and approved_count == total_count:
                color = (80, 250, 123) # Green
            elif approved_count > 0:
                color = (255, 184, 108) # Orange
            else:
                color = (100, 100, 100) # Gray
                
            self.stats_pill.setStyleSheet(f"""
                background-color: rgba(74, 158, 255, 0.08);
                color: #4a9eff;
                padding: 1px 8px;
                border-radius: 10px;
                font-size: 10px;
                border: 1px solid rgba(74, 158, 255, 0.2);
            """)
        elif hasattr(self, 'stats_label'):
            # Для Minimalist
            self.stats_label.setText(f"✅ {approved_count} | 📦 {used_count} | 📝 {total_count}")


class GeneratedReviewsTab(QWidget):
    """Вкладка с календарным просмотром отзывов."""
    
    def __init__(self):
        super().__init__()
        self.current_period_id = None
        self.period_start = None
        self.period_end = None
        self.view_mode = "dashboard" # dashboard, kanban, minimalist
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Статистика и инфо
        self.stats_label = QLabel("Выберите период для просмотра отзывов")
        self.stats_label.setStyleSheet("font-size: 14px; padding: 10px; background-color: #2d2d2d; border-radius: 5px;")
        layout.addWidget(self.stats_label)
        
        # Информация о товарах без отзывов
        self.missing_reviews_text = QTextEdit()
        self.missing_reviews_text.setStyleSheet("""
            QTextEdit {
                font-size: 11px; 
                padding: 8px; 
                background-color: rgba(74, 158, 255, 0.12);
                color: #cdd6f4;
                border-radius: 8px;
                margin: 5px 0;
                border: 1px solid rgba(74, 158, 255, 0.35);
            }
        """)
        self.missing_reviews_text.setFixedHeight(80)  # Фиксированная высота
        self.missing_reviews_text.setReadOnly(True)  # Только для чтения, но с возможностью копирования
        self.missing_reviews_text.hide()  # Скрыто по умолчанию
        layout.addWidget(self.missing_reviews_text)
        
        # Тулбар
        toolbar = QHBoxLayout()
        
        # Поиск
        search_label = QLabel("Поиск:")
        search_label.setStyleSheet("color: #ccc; margin-right: 5px;")
        toolbar.addWidget(search_label)
        
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Найти отзыв...")
        self.search_edit.setStyleSheet("""
            QLineEdit {
                background-color: #333;
                color: #fff;
                border: 1px solid #555;
                border-radius: 4px;
                padding: 5px;
                min-width: 200px;
            }
            QLineEdit:focus {
                border-color: #4a9eff;
            }
        """)
        self.search_edit.textChanged.connect(self.filter_reviews)
        toolbar.addWidget(self.search_edit)
        
        toolbar.addStretch()
        
        # Neon кнопки с цветовой схемой "Лед и Пламя"
        self.distribute_btn = NeonButton("Распределить по датам", "primary")  # Ледяной синий
        self.distribute_btn.clicked.connect(self.distribute_reviews_action)
        toolbar.addWidget(self.distribute_btn)

        # Переключатель видов (Switcher)
        toolbar.addSpacing(20)
        mode_layout = QHBoxLayout()
        mode_layout.setSpacing(2)
        
        self.mode_dash_btn = NeonButton("📊", "primary")  # Set as primary by default
        self.mode_dash_btn.setToolTip("Дашборд (Полная информация)")
        self.mode_dash_btn.setFixedWidth(40)
        self.mode_dash_btn.clicked.connect(lambda: self.switch_view("dashboard"))
        
        self.mode_kanban_btn = NeonButton("📋", "secondary")
        self.mode_kanban_btn.setToolTip("Канбан (Компактно с деталями)")
        self.mode_kanban_btn.setFixedWidth(40)
        self.mode_kanban_btn.clicked.connect(lambda: self.switch_view("kanban"))
        
        self.mode_minimal_btn = NeonButton("🌙", "secondary")
        self.mode_minimal_btn.setToolTip("Минимализм (Только даты)")
        self.mode_minimal_btn.setFixedWidth(40)
        self.mode_minimal_btn.clicked.connect(lambda: self.switch_view("minimalist"))
        
        mode_layout.addWidget(self.mode_dash_btn)
        mode_layout.addWidget(self.mode_kanban_btn)
        mode_layout.addWidget(self.mode_minimal_btn)
        toolbar.addLayout(mode_layout)

        toolbar.addStretch()

        self.clear_btn = NeonButton("Очистить все", "secondary")  # Оранжевый
        self.clear_btn.clicked.connect(self.clear_all)
        self.clear_btn.setEnabled(False)
        toolbar.addWidget(self.clear_btn)
        
        toolbar.addStretch()
        
        self.export_btn = NeonButton("Экспорт Excel", "suggested")  # Золотой - рекомендуемое
        self.export_btn.clicked.connect(self.export_excel)
        self.export_btn.setEnabled(False)
        toolbar.addWidget(self.export_btn)
        
        self.delete_reviews_btn = NeonButton("Удалить отзывы", "danger")
        self.delete_reviews_btn.clicked.connect(self.delete_reviews)
        self.delete_reviews_btn.setEnabled(False)
        toolbar.addWidget(self.delete_reviews_btn)
        
        self.generate_missing_btn = NeonButton("Генерировать недостающие", "secondary")
        self.generate_missing_btn.clicked.connect(self.generate_missing_reviews)
        self.generate_missing_btn.setEnabled(False)
        self.generate_missing_btn.hide()  # Скрыта по умолчанию
        toolbar.addWidget(self.generate_missing_btn)
        
        layout.addLayout(toolbar)
        
        # Scroll area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        
        # Контейнер для списка дней
        self.days_container = QWidget()
        # Все режимы используют сетку 7 колонок с разными отступами
        self.days_layout = QGridLayout() 
        if self.view_mode == "dashboard":
            self.days_layout.setSpacing(10)
        elif self.view_mode == "kanban":
            self.days_layout.setSpacing(6)
        else:  # minimalist
            self.days_layout.setSpacing(4)
        self.days_layout.setContentsMargins(10, 10, 10, 10)
        self.days_container.setLayout(self.days_layout)
        
        scroll.setWidget(self.days_container)
        layout.addWidget(scroll)
        
        self.setLayout(layout)
    
    def switch_view(self, mode):
        """Переключить режим отображения календаря."""
        print(f"DEBUG: switch_view CALLED with mode: {mode}")
        if self.view_mode == mode:
            print(f"DEBUG: mode is already {mode}, skipping.")
            return
            
        # Сначала очищаем все старые виджеты, пока они в старой сетке
        while self.days_layout.count():
            item = self.days_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
            
        self.view_mode = mode
        # Сброс стилей кнопок
        for btn in [self.mode_dash_btn, self.mode_kanban_btn, self.mode_minimal_btn]:
            btn.set_button_type("secondary")
            
        if mode == "dashboard": self.mode_dash_btn.set_button_type("primary")
        elif mode == "kanban": self.mode_kanban_btn.set_button_type("primary")
        elif mode == "minimalist": self.mode_minimal_btn.set_button_type("primary")
        
        # Пересоздаем лейаут
        QWidget().setLayout(self.days_container.layout())
        
        if mode == "dashboard":
            new_layout = QGridLayout()
            new_layout.setSpacing(10)
        elif mode == "kanban":
            new_layout = QGridLayout()
            new_layout.setSpacing(6)
        else:  # minimalist
            new_layout = QGridLayout()
            new_layout.setSpacing(4)
            
        new_layout.setContentsMargins(10, 10, 10, 10)
        new_layout.setAlignment(Qt.AlignmentFlag.AlignTop) # ПРЕДОТВРАЩАЕМ РАСТЯГИВАНИЕ
        self.days_container.setLayout(new_layout)
        self.days_layout = new_layout
        
        print(f"DEBUG: Layout switched. self.view_mode is now: {self.view_mode}")
        self.load_reviews()

    def open_review_dialog(self, review_card):
        """Открыть диалог редактирования отзыва."""
        from ui.dialogs.review_edit_dialog import ReviewEditDialog
        from core.database import db
        from core.models import Review
        
        with db.get_session() as session:
            review = session.query(Review).get(review_card.review_id)
            if review:
                dialog = ReviewEditDialog(review, parent=self)
                # Перезагружаем интерфейс после любого изменения (сохранение или удаление)
                dialog.exec()
                self.load_reviews()

    def distribute_reviews_action(self):
        """Ручное распределение отзывов по датам."""
        from core.database import db
        from core.models import Review
        with db.get_session() as session:
            reviews = session.query(Review).filter_by(
                period_id=self.current_period_id,
                is_generated=True
            ).all()
            if reviews:
                self._distribute_reviews_logic(session, reviews, self.period_start, self.period_end)
                self.load_reviews()
    
    def set_period(self, period_id):
        """Установить период для просмотра."""
        from core.database import db
        from core.models import Period
        
        with db.get_session() as session:
            period = session.query(Period).get(period_id)
            if period:
                self.current_period_id = period_id
                self.period_start = period.start_date
                self.period_end = period.end_date
                
                self.clear_btn.setEnabled(True)
                
                # Проверяем условия для активации экспорта
                self._update_export_button_state()
                
                # Проверяем товары без отзывов и с недостаточным количеством
                self._check_missing_reviews()
                
                self.load_reviews()
    
    def _check_missing_reviews(self):
        """Проверить товары без отзывов и показать информацию."""
        if not self.current_period_id:
            self.missing_reviews_text.hide()
            return
        
        from core.database import db
        from core.models import ProductTask, Review
        
        with db.get_session() as session:
            # Получаем все товары в периоде
            products = session.query(ProductTask).filter_by(
                period_id=self.current_period_id
            ).all()
            
            if not products:
                self.missing_reviews_text.hide()
                return
            
            # Получаем товары с отзывами
            products_with_reviews = session.query(Review.product_task_id).filter(
                Review.period_id == self.current_period_id
            ).distinct().all()
            products_with_reviews = set(p[0] for p in products_with_reviews)
            
            # Находим товары без отзывов и с недостаточным количеством
            missing_products = []
            insufficient_products = []
            
            for product in products:
                if product.id not in products_with_reviews:
                    missing_products.append(product)
                elif product.review_count and product.review_count > 0:
                    # Проверяем количество сгенерированных отзывов
                    generated_count = session.query(Review).filter(
                        Review.product_task_id == product.id,
                        Review.is_generated == True
                    ).count()
                    
                    if generated_count < product.review_count:
                        insufficient_products.append(product)
            
            # Формируем текст сообщения
            text_parts = []
            
            if missing_products:
                text_parts.append("⚠️ Товары без отзывов:")
                for i, product in enumerate(missing_products[:5], 1):
                    text_parts.append(f"{i}. {product.product_name}")
                    if product.product_url:
                        text_parts.append(f"   URL: {product.product_url}")
                
                if len(missing_products) > 5:
                    text_parts.append(f"... и еще {len(missing_products) - 5} товаров")
                text_parts.append("")  # Пустая строка
            
            if insufficient_products:
                text_parts.append("⚠️ Недостаточно отзывов:")
                for i, product in enumerate(insufficient_products[:5], 1):
                    text_parts.append(f"{i}. {product.product_name} (нужно {product.review_count}, сгенерировано {session.query(Review).filter(Review.product_task_id == product.id, Review.is_generated == True).count()})")
                
                if len(insufficient_products) > 5:
                    text_parts.append(f"... и еще {len(insufficient_products) - 5} товаров")
            
            if text_parts:
                self.missing_reviews_text.setPlainText("\n".join(text_parts))
                self.missing_reviews_text.show()
                
                # Показываем кнопку генерации недостающих
                self.generate_missing_btn.show()
                self.generate_missing_btn.setEnabled(True)
                # Обновляем текст кнопки
                total_missing = len(missing_products) + len(insufficient_products)
                self.generate_missing_btn.setText(f"Генерировать недостающие ({total_missing})")
            else:
                self.missing_reviews_text.hide()
                self.generate_missing_btn.hide()
                self.generate_missing_btn.setEnabled(False)
    
    def load_reviews(self):
        """Загрузить отзывы и отрисовать календарь."""
        # Очистка текущего лейаута
        while self.days_layout.count():
            item = self.days_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        if not self.current_period_id or not self.period_start or not self.period_end:
            return

        from core.database import db
        from core.models import Review
        from datetime import timedelta
        
        with db.get_session() as session:
            # Получить все сгенерированные отзывы
            reviews = session.query(Review).filter_by(
                period_id=self.current_period_id,
                is_generated=True
            ).all()
            
            # Предварительное распределение (только если есть нераспределённые)
            reviews_without_date = [r for r in reviews if not r.target_date]
            if reviews_without_date:
                self._distribute_reviews_logic(session, reviews_without_date, self.period_start, self.period_end)
                # Повторный запрос только если были распределения
                reviews = session.query(Review).filter_by(
                    period_id=self.current_period_id,
                    is_generated=True
                ).order_by(Review.target_date).all()
            
            self.stats_label.setText(f"<b>Сгенерировано отзывов:</b> {len(reviews)}")
            self._check_missing_reviews()
            
            # Группировка отзывов по датам
            reviews_by_date = {}
            for review in reviews:
                if review.target_date:
                    d = review.target_date.date() if hasattr(review.target_date, 'date') else review.target_date
                    if d not in reviews_by_date: 
                        reviews_by_date[d] = []
                    reviews_by_date[d].append(review)
            
            # Отрисовка дней
            current_date = self.period_start
            index = 0
            while current_date <= self.period_end:
                # Преобразуем в date для сравнения
                current_date_key = current_date.date() if hasattr(current_date, 'date') else current_date
                day_reviews = reviews_by_date.get(current_date_key, [])
                
                # Создаем контейнер дня
                day_widget = DayContainer(current_date, view_mode=self.view_mode, parent=self)
                
                # Добавляем в лейаут
                # Все режимы теперь используют сетку 7 колонок
                row = index // 7
                col = index % 7
                self.days_layout.addWidget(day_widget, row, col)
                
                # Считаем статистику для дня
                approved = sum(1 for r in day_reviews if r.is_approved)
                used = sum(1 for r in day_reviews if r.is_used)
                day_widget.update_stats(approved, used, len(day_reviews))
                
                # Добавляем отзывы
                if self.view_mode != "minimalist":
                    for review in day_reviews:
                        card = ReviewCard(review, view_mode=self.view_mode, parent=day_widget)
                        card.approved_changed.connect(self.update_review_status)
                        card.used_changed.connect(self.update_review_used)
                        card.open_btn.clicked.connect(lambda ch, c=card: self.open_review_dialog(c))
                        day_widget.add_review(card)
                
                current_date += timedelta(days=1)
                index += 1
    
    def filter_reviews(self, text):
        """Фильтрация отзывов по поисковому тексту."""
        search_text = text.lower().strip()
        
        # Проходим по всем дням
        for i in range(self.days_layout.count() - 1):  # -1 чтобы не считать stretch
            item = self.days_layout.itemAt(i)
            if item and item.widget():
                day_container = item.widget()
                if hasattr(day_container, 'reviews_area'):
                    show_day = False
                    
                    # Проверяем каждый отзыв в дне
                    for j in range(day_container.reviews_area.count()):
                        review_item = day_container.reviews_area.itemAt(j)
                        if review_item and review_item.widget():
                            review_card = review_item.widget()
                            if hasattr(review_card, 'review_id'):
                                # Ищем по сохранённым данным
                                review_text = f"{review_card.review_id} {review_card.product_name} {review_card.author} {review_card.content}".lower()
                                if search_text in review_text:
                                    review_card.show()
                                    show_day = True
                                else:
                                    review_card.hide()
                    
                    # Показываем/скрываем весь день
                    day_container.setVisible(show_day or not search_text)
    
    def delete_reviews(self):
        """Удаление всех отзывов для текущего периода."""
        if not self.current_period_id:
            return
        
        reply = QMessageBox.question(
            self, 
            "Подтверждение", 
            "Удалить все отзывы для этого периода?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            from core.database import db
            from core.models import Review
            
            with db.get_session() as session:
                deleted = session.query(Review).filter(
                    Review.period_id == self.current_period_id
                ).delete()
                session.commit()
                
                QMessageBox.information(
                    self, 
                    "Успех", 
                    f"Удалено {deleted} отзывов"
                )
                
                # Перезагрузить отзывы
                self.load_reviews()



    def distribute_reviews_action(self):
        """Ручной вызов перераспределения."""
        if not self.current_period_id: 
            return
            
        from core.database import db
        from core.models import Review
        
        with db.get_session() as session:
            reviews = session.query(Review).filter_by(
                period_id=self.current_period_id,
                is_generated=True
            ).all()
            
            self._distribute_reviews_logic(session, reviews, self.period_start, self.period_end)
            
        self.load_reviews()

    def _distribute_reviews_logic(self, session, reviews, start_date, end_date):
        """Логика распределения отзывов по датам."""
        if not reviews:
            return

        from datetime import timedelta
        total_days = (end_date - start_date).days + 1
        if total_days <= 0: total_days = 1
        
        # Простая логика: равномерно размазать
        # Можно улучшить: "схожие отзывы подальше" (пока просто round-robin)
        
        import random
        
        # Преобразуем start_date в datetime если надо
        curr = start_date
        
        # Шаг распределения: если отзывов меньше чем дней, кидаем случайно или равномерно
        # Если отзывов больше - заполняем каждый день
        
        # Создаем список всех доступных дат
        all_dates = [start_date + timedelta(days=i) for i in range(total_days)]
        
        # Перемешиваем отзывы чтобы схожие (которые часто идут подряд при генерации) встали в разные места
        # Но пользователь просил "схожие подальше". Если мы их генерили пачкой для одного товара, они мб подряд.
        # Random shuffle неплохо справляется с базовым разбросом.
        reviews_to_update = list(reviews)
        random.shuffle(reviews_to_update)
        
        # Распределяем
        for i, review in enumerate(reviews_to_update):
            # i % total_days дает индекс даты. Это обеспечит равномерное заполнение.
            date_idx = i % total_days
            target_date = all_dates[date_idx]
            
            review.target_date = target_date
            
        session.commit()

    
    def update_review_status(self, review_id, is_approved):
        """Обработать изменение чекбокса ОК с обработкой ошибок."""
        try:
            from core.database import db
            from core.models import Review
            
            with db.get_session() as session:
                review = session.query(Review).get(review_id)
                if review:
                    review.is_approved = is_approved
                    session.commit()
                else:
                    print(f"Ошибка: Отзыв с ID {review_id} не найден")
            
            # Обновляем состояние кнопки экспорта
            self._update_export_button_state()
        except Exception as e:
            print(f"Ошибка при изменении статуса утверждения: {e}")
    
    def update_review_used(self, review_id, is_used):
        """Обработать изменение чекбокса Использовано с обработкой ошибок."""
        try:
            from core.database import db
            from core.models import Review
            
            with db.get_session() as session:
                review = session.query(Review).get(review_id)
                if review:
                    review.is_used = is_used
                    session.commit()
                else:
                    print(f"Ошибка: Отзыв с ID {review_id} не найден")
            
            # Обновляем состояние кнопки экспорта
            self._update_export_button_state()
        except Exception as e:
            print(f"Ошибка при изменении статуса использования: {e}")
    
    def clear_all(self):
        """Очистить все отзывы периода."""
        if not self.current_period_id:
            return
        
        reply = QMessageBox.question(
            self, "Подтверждение",
            "Удалить ВСЕ сгенерированные отзывы этого периода?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            from core.database import db
            from core.models import Review
            
            with db.get_session() as session:
                deleted = session.query(Review).filter_by(
                    period_id=self.current_period_id,
                    is_generated=True
                ).delete()
                session.commit()
            
            QMessageBox.information(self, "Успех", f"Удалено {deleted} отзывов")
            self.load_reviews()
    
    def _update_export_button_state(self):
        """Обновить состояние кнопки экспорта."""
        if not self.current_period_id:
            self.export_btn.setEnabled(False)
            return
        
        from core.database import db
        from core.models import Review
        
        with db.get_session() as session:
            # Проверяем все отзывы для периода
            total_reviews = session.query(Review).filter_by(
                period_id=self.current_period_id,
                is_generated=True
            ).count()
            
            if total_reviews == 0:
                self.export_btn.setEnabled(False)
                self.export_btn.setToolTip("Нет отзывов для экспорта")
                return
            
            # Проверяем что все отзывы приняты
            approved_reviews = session.query(Review).filter_by(
                period_id=self.current_period_id,
                is_generated=True,
                is_approved=True
            ).count()
            
            # Проверяем что все отзывы использованы
            used_reviews = session.query(Review).filter_by(
                period_id=self.current_period_id,
                is_generated=True,
                is_used=True
            ).count()
            
            # Активируем кнопку только если все отзывы приняты и использованы
            can_export = (approved_reviews == total_reviews and used_reviews == total_reviews)
            
            self.export_btn.setEnabled(can_export)
            
            if can_export:
                self.export_btn.setToolTip(f"Экспорт {total_reviews} отзывов в Excel")
            else:
                remaining_approved = total_reviews - approved_reviews
                remaining_used = total_reviews - used_reviews
                self.export_btn.setToolTip(
                    f"Нужно принять: {remaining_approved}, использовать: {remaining_used}"
                )
    
    def export_excel(self):
        """Экспорт отзывов в Excel."""
        if not self.current_period_id:
            QMessageBox.warning(self, "Ошибка", "Период не выбран")
            return
        
        from core.database import db
        from core.models import Review, Period
        from PyQt6.QtWidgets import QFileDialog
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        try:
            with db.get_session() as session:
                # Получаем информацию о периоде
                period = session.query(Period).get(self.current_period_id)
                if not period:
                    QMessageBox.warning(self, "Ошибка", "Период не найден")
                    return
                
                # Получаем все принятые и использованные отзывы
                reviews = session.query(Review).filter_by(
                    period_id=self.current_period_id,
                    is_generated=True,
                    is_approved=True,
                    is_used=True
                ).order_by(Review.target_date).all()
                
                if not reviews:
                    QMessageBox.warning(self, "Ошибка", "Нет отзывов для экспорта")
                    return
                
                # Выбираем файл для сохранения
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Сохранить Excel файл",
                    f"отзывы_{period.start_date.strftime('%Y-%m-%d')}.xlsx",
                    "Excel Files (*.xlsx)"
                )
                
                if not file_path:
                    return
                
                # Создаем Excel файл
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Отзывы"
                
                # Заголовки
                headers = [
                    "ID", "Товар", "Автор", "Рейтинг", 
                    "Отзыв", "Плюсы", "Минусы", "Дата", "Источник", "URL размещения"
                ]
                
                # Стили для заголовков
                header_font = Font(bold=True, size=12)
                header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # Записываем заголовки
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Записываем отзывы
                for row, review in enumerate(reviews, 2):
                    ws.cell(row=row, column=1, value=review.id)
                    ws.cell(row=row, column=2, value=review.product_name or "")
                    ws.cell(row=row, column=3, value=review.author or "")
                    ws.cell(row=row, column=4, value=review.rating or "")
                    ws.cell(row=row, column=5, value=review.content or "")
                    ws.cell(row=row, column=6, value=review.pros or "")
                    ws.cell(row=row, column=7, value=review.cons or "")
                    ws.cell(row=row, column=8, value=review.target_date.strftime("%d.%m.%Y") if review.target_date else "")
                    ws.cell(row=row, column=9, value=review.source or "")
                    ws.cell(row=row, column=10, value=review.placement_url or "")
                
                # Автоматическая ширина колонок
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Сохраняем файл
                wb.save(file_path)
                
                QMessageBox.information(
                    self, 
                    "Успех", 
                    f"Экспортировано {len(reviews)} отзывов в файл:\n{file_path}"
                )
                
        except Exception as e:
            QMessageBox.critical(
                self, 
                "Ошибка экспорта", 
                f"Произошла ошибка при экспорте:\n{str(e)}"
            )
    
    def generate_missing_reviews(self):
        """Генерировать отзывы для товаров без отзывов."""
        if not self.current_period_id:
            QMessageBox.warning(self, "Ошибка", "Период не выбран")
            return
        
        try:
            from core.database import db
            from core.models import ProductTask, Review
            
            with db.get_session() as session:
                # Получаем товары без отзывов
                products = session.query(ProductTask).filter_by(
                    period_id=self.current_period_id
                ).all()
                
                products_with_reviews = session.query(Review.product_task_id).filter_by(
                    period_id=self.current_period_id,
                    is_generated=True
                ).distinct().all()
                
                products_with_reviews_ids = [p[0] for p in products_with_reviews]
                missing_products = [p for p in products if p.id not in products_with_reviews_ids]
                
                if not missing_products:
                    QMessageBox.information(self, "Информация", "Все товары имеют отзывы")
                    return
                
                # Подтверждение
                product_names = [p.product_name for p in missing_products[:5]]
                text = f"Сгенерировать отзывы для {len(missing_products)} товаров?\n\n"
                text += "Товары:\n" + "\n".join([f"• {name}" for name in product_names])
                
                if len(missing_products) > 5:
                    text += f"\n• ... и еще {len(missing_products) - 5} товаров"
                
                reply = QMessageBox.question(
                    self, 
                    "Подтверждение", 
                    text,
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                
                if reply != QMessageBox.StandardButton.Yes:
                    return
                
                # Запускаем генерацию
                from ui.dialogs.generate_dialog import GenerateDialog
                product_ids = [p.id for p in missing_products]
                
                dialog = GenerateDialog(parent=self, product_ids=product_ids)
                if dialog.exec():
                    # Обновляем после генерации
                    self.load_reviews()
                    QMessageBox.information(
                        self, 
                        "Успех", 
                        f"Сгенерировано отзывы для {len(missing_products)} товаров"
                    )
                    
        except Exception as e:
            QMessageBox.critical(
                self, 
                "Ошибка генерации", 
                f"Произошла ошибка при генерации:\n{str(e)}"
            )
